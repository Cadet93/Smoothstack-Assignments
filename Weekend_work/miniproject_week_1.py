import math
import openpyxl as pyxl
from openpyxl.styles import numbers
import datetime
from datetime import datetime as dt
import logging
import os

#ensure working directory is correct
os.chdir('.\\Smoothstack-Assignments\\Weekend_work')

filename = "expedia_report_monthly_january_2018.xlsx"
log_filename = "logfile.log"

logging.basicConfig(format='%(asctime)s %(message)s',
                datefmt='%m/%d/%Y %I:%M:%S %p',
                filename=log_filename,
                filemode='w',
                level=logging.INFO)

logging.info('miniproject Started...')

def get_year_month(filename):
    try:
        #create variables for year and month of file
        year = filename[-9:-5]

        month = filename.split("_")[-2]
        
        #convert month into month number
        month_number = datetime.datetime.strptime(month, '%B').month

        #determine if month number is 1 or 2 digits
        mon_length = int(math.log10(month_number))+1
    
        #ensure month is 2 digits for search string
        if mon_length == 1:
            month_number = '0' + str(month_number)
        else:
            month_number = str(month_number)

        #create search string year and month
        year_month = year + "-" + month_number

        return(year_month)
    except:
        logging.error("file name in incorrect format.")

def summary(wb, year_month):
    
    #set correct sheet
    sheet = sheet = wb['Summary Rolling MoM']

    #set header boolean and current row variables
    header = True
    current_row = 1
    
    #loop through rows until getting to correct row for year and month of file
    for row in sheet.rows:

        #skip header row and change boolean
        if header:
            header = False
            current_row +=1
            continue

        #grab datetime from current row
        row_date_time = str(row[0].value)
        
        #strip datetime to only year and month
        row_year_month = '{:%Y-%m}'.format(dt.strptime(row_date_time, '%Y-%m-%d %H:%M:%S'))
        
        #if file year and month match row year and month print values to log file
        if row_year_month == year_month:
            
            #log values from summary tab
            co= '{:,}'.format(sheet.cell(current_row, column=2).value)
            logging.info(f"{sheet.cell(row=1, column=2).value}: {co}")
            
            abandon= "{:.2%}".format(sheet.cell(current_row, column=3).value)
            logging.info(f"{sheet.cell(row=1, column=3).value}: {abandon}")
            
            fcr= "{:.2%}".format(sheet.cell(current_row, column=4).value)
            logging.info(f"{sheet.cell(row=1, column=4).value}: {fcr}")
            
            dsat= "{:.2%}".format(sheet.cell(current_row, column=5).value)
            logging.info(f"{sheet.cell(row=1, column=5).value}: {dsat}")
            
            csat= "{:.2%}".format(sheet.cell(current_row, column=6).value)
            logging.info(f"{sheet.cell(row=1, column=6).value}: {csat}")

            #end loop once correct month is found.
            break

        #increment row counter before moving to next row
        current_row +=1

def voc(wb, filename):

    #switch to the VOC Rolling Mom tab in the workbook
    sheet = wb['VOC Rolling MoM']
    
    #set header boolean 
    header = True
    current_col = 1
    
    #loop through the columns to find the month and year of the file
    for col in sheet.iter_cols(min_row=1, max_row=1):
        
        #skip header row and change boolean
        if header:
            header = False
            current_col +=1
            continue  
            
        #grab datetime from current row
        col_date_time = str(col[0].value)
        
        #strip datetime to only year and month
        col_year_month = '{:%Y-%m}'.format(dt.strptime(col_date_time, '%Y-%m-%d %H:%M:%S'))
        
        
        #if file year and month match row year and month print values to log file
        if col_year_month == year_month:
            
            #log NPS header
            logging.info(sheet.cell(row=2, column=1).value)
            
            #log base size
            bs= '{:,}'.format(sheet.cell(row=3, column=current_col).value)
            logging.info(f"{sheet.cell(row=3, column=1).value}: {bs}")
            
            #grab promoter number and determine if its good or bad
            promoters = sheet.cell(row=4, column=current_col).value
            
            logging.info(f"Number of Promoters: {promoters}")
            
            if int(promoters) >= 200:
                promoters_result = 'good'
            else:
                promoters_result = 'bad'
                
            logging.info(f"{sheet.cell(row=4, column=1).value}: {promoters_result}")
            
            #grab passive number and determine if its good or bad
            passives = sheet.cell(row=6, column=current_col).value
            
            logging.info(f"Number of Passives: {passives}")
            
            if int(passives) >= 100:
                passives_result = 'good'
            else:
                passives_result = 'bad'
                
            logging.info(f"{sheet.cell(row=6, column=1).value}: {passives_result}")
            
            #grab detractors number and determine if its good or bad
            detractors = sheet.cell(row=8, column=current_col).value
            
            logging.info(f"Number of Detractors: {detractors}")
            
            if int(detractors) <= 100:
                detractors_result = 'good'
            else:
                detractors_result = 'bad'
                
            logging.info(f"{sheet.cell(row=8, column=1).value}: {detractors_result}")
            
            #log NPS info
            logging.info(sheet.cell(row=12, column=1).value)
            nps ="{:.2%}".format(sheet.cell(row=13, column = current_col).value)
            logging.info(f"{sheet.cell(row=13, column=1).value}: {nps}")
            logging.info(sheet.cell(row=15, column=1).value)
            sat ="{:.2%}".format(sheet.cell(row=16, column = current_col).value)
            logging.info(f"{sheet.cell(row=16, column=1).value}: {sat}")
            logging.info(sheet.cell(row=15, column=1).value)
            dsat ="{:.2%}".format(sheet.cell(row=19, column = current_col).value)
            logging.info(f"{sheet.cell(row=19, column=1).value}: {dsat}")
            
            #end loop once correct month is found.
            break
            
        #increment row counter before moving to next row
        current_col +=1

#call function to strip year and month number from file provided
year_month = get_year_month(filename)

#exception handling for file errors
try:
    #open file to current sheet to log needed data from the summary tab
    wb = pyxl.open(filename)

    #call function to log data from the summary tab
    summary(wb, year_month)

    #call function to log data from the VOC tab
    voc(wb, filename)    
except OSError as e:
    #log error
    logging.error(e)




